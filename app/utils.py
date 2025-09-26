import qrcode
import json
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from flask import current_app, make_response
from datetime import datetime

def generate_qr_code(item):
    """Generate QR code for an item with current details"""
    qr_data = json.dumps(item.get_qr_data(), indent=2)
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to file
    filename = f"qr_{item.ssid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    img.save(filepath)
    
    return filename

def create_excel_response(data, filename, headers):
    """Create Excel file and return as response"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add data
    for row, record in enumerate(data, 2):
        for col, value in enumerate(record, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response
